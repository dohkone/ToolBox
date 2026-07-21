#!/usr/bin/env python3
"""Generate prompts or images from the local JSON/Excel template library."""

from __future__ import annotations

import argparse
import concurrent.futures
import json
import random
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ModuleNotFoundError:
    openpyxl = None

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


DEFAULT_TEMPLATE_PATH = Path(r"D:\temu_auto\temp\文生图模板库_Codex.xlsx")
DEFAULT_OUTPUT_DIR = Path(r"D:\temu_auto\review")
SHEET_NAME = "文生图模板库"
LAYOUT_COLUMN = "布局模板"
SCENE_COLUMN = "场景模板"
SUBJECT_COLUMN = "主体"

COLOR_OPTIONS = (
    ("黑色", "#0A0A0A"),
    ("米白色", "#F4F4F2"),
    ("深棕色", "#261107"),
    ("深灰色", "#C4C8CA"),
    ("酒红色", "#722829"),
    ("宝蓝色", "#2E3EA5"),
)

COLOR_PLACEHOLDER = "{颜色}"
SCENE_PLACEHOLDER = "{场景}"
SUBJECT_PLACEHOLDER = "{主体}"
FIXED_COLOR_PLACEHOLDER = "{唯一颜色}"
FIXED_SUBJECT_PLACEHOLDER = "{唯一主体}"
ALL_COLORS_PLACEHOLDER = "{全部颜色}"
ALL_SUBJECTS_PLACEHOLDER = "{全部主体}"


class TemplateRandomError(Exception):
    """Raised when prompt generation fails."""


@dataclass
class SceneTemplateEntry:
    content: str
    subjects: list[str]


@dataclass
class SelectedTemplate:
    layout_template: str
    scene_template: str
    subject: str


@dataclass
class TemplateLibrary:
    layout_templates: list[str]
    scene_templates: list[SceneTemplateEntry]
    subject_templates: list[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate prompts from the local template library.")
    parser.add_argument("--template-path", default=str(DEFAULT_TEMPLATE_PATH))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--count", type=int, default=5)
    parser.add_argument("--concurrency", type=int, default=2)
    parser.add_argument("--image2-script")
    parser.add_argument("--seed", type=int)
    parser.add_argument("--unique-scene", action="store_true")
    parser.add_argument("--prompts-only", action="store_true")
    return parser.parse_args()


def split_variants(value: str | None) -> list[str]:
    if value is None:
        return []

    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    parts = re.split(r"[\/\n]+", text)
    return [part.strip(" -\t") for part in parts if part and part.strip(" -\t")]


def split_variants_list(values: list[str] | tuple[str, ...]) -> list[str]:
    result: list[str] = []
    for value in values:
        result.extend(split_variants(value))
    return result


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        key = value.strip()
        if not key:
            continue
        lowered = key.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        result.append(key)
    return result


def dedupe_scene_entries(entries: list[SceneTemplateEntry]) -> list[SceneTemplateEntry]:
    seen: set[tuple[str, tuple[str, ...]]] = set()
    result: list[SceneTemplateEntry] = []
    for entry in entries:
        content = entry.content.strip()
        if not content:
            continue
        subjects = tuple(dedupe_preserve_order(entry.subjects))
        key = (content.lower(), tuple(subject.lower() for subject in subjects))
        if key in seen:
            continue
        seen.add(key)
        result.append(SceneTemplateEntry(content=content, subjects=list(subjects)))
    return result


def load_template_library(template_path: Path) -> TemplateLibrary:
    if not template_path.exists():
        raise TemplateRandomError(f"模板库不存在：{template_path}")

    if template_path.suffix.lower() == ".json":
        return load_template_library_from_json(template_path)

    if openpyxl is None:
        raise TemplateRandomError("当前模板库是 Excel，但运行环境缺少 openpyxl。")

    return load_template_library_from_excel(template_path)


def load_template_library_from_json(template_path: Path) -> TemplateLibrary:
    payload = json.loads(template_path.read_text(encoding="utf-8-sig"))

    is_exported_generation_library = any(
        key in payload for key in ("LayoutTemplates", "SceneTemplates", "SubjectTemplates")
    )

    layout_values = payload.get("layout_templates", payload.get("LayoutTemplates", []))
    scene_values = payload.get("scene_templates", payload.get("SceneTemplates", []))
    subject_values = payload.get("subject_templates", payload.get("SubjectTemplates", []))

    if is_exported_generation_library:
        layout_templates = dedupe_preserve_order(
            [str(value).strip() for value in layout_values if str(value).strip()]
        )
        subject_templates = dedupe_preserve_order(
            [str(value).strip() for value in subject_values if str(value).strip()]
        )
    else:
        layout_templates = dedupe_preserve_order(split_variants_list(layout_values))
        subject_templates = dedupe_preserve_order(split_variants_list(subject_values))

    scene_entries: list[SceneTemplateEntry] = []
    for raw_scene in scene_values:
        if isinstance(raw_scene, str):
            for content in split_variants(raw_scene):
                scene_entries.append(SceneTemplateEntry(content=content, subjects=[]))
            continue

        if not isinstance(raw_scene, dict):
            continue

        contents = split_variants(str(raw_scene.get("content", raw_scene.get("Content", ""))))
        raw_subjects = raw_scene.get("subjects", raw_scene.get("Subjects", []))
        if isinstance(raw_subjects, list):
            if is_exported_generation_library:
                subjects = dedupe_preserve_order(
                    [str(item).strip() for item in raw_subjects if str(item).strip()]
                )
            else:
                subjects = dedupe_preserve_order(split_variants_list([str(item) for item in raw_subjects]))
        else:
            subjects = dedupe_preserve_order(split_variants(str(raw_subjects)))

        for content in contents:
            scene_entries.append(SceneTemplateEntry(content=content, subjects=subjects))

    return build_library(layout_templates, scene_entries, subject_templates)


def load_template_library_from_excel(template_path: Path) -> TemplateLibrary:
    workbook = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise TemplateRandomError(f"模板库中缺少工作表：{SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise TemplateRandomError("模板库没有可用数据。")

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    try:
        layout_index = header.index(LAYOUT_COLUMN)
        scene_index = header.index(SCENE_COLUMN)
        subject_index = header.index(SUBJECT_COLUMN)
    except ValueError as exc:
        raise TemplateRandomError("模板库表头缺少“布局模板 / 场景模板 / 主体”列。") from exc

    layout_templates: list[str] = []
    scene_entries: list[SceneTemplateEntry] = []
    subject_templates: list[str] = []

    for raw_row in rows[1:]:
        if not raw_row:
            continue

        layout = raw_row[layout_index] if len(raw_row) > layout_index else None
        scene = raw_row[scene_index] if len(raw_row) > scene_index else None
        subject = raw_row[subject_index] if len(raw_row) > subject_index else None

        if layout and str(layout).strip():
            layout_templates.append(str(layout).strip())

        scene_variants = split_variants(scene)
        subject_variants = dedupe_preserve_order(split_variants(subject))
        subject_templates.extend(subject_variants)

        for scene_variant in scene_variants:
            scene_entries.append(SceneTemplateEntry(content=scene_variant, subjects=subject_variants))

    return build_library(layout_templates, scene_entries, subject_templates)


def build_library(
    layout_templates: list[str],
    scene_entries: list[SceneTemplateEntry],
    subject_templates: list[str],
) -> TemplateLibrary:
    layouts = dedupe_preserve_order(layout_templates)
    scenes = dedupe_scene_entries(scene_entries)
    subjects = dedupe_preserve_order(subject_templates)

    if not layouts:
        raise TemplateRandomError("模板库里没有可用的布局模板。")
    if not scenes:
        raise TemplateRandomError("模板库里没有可用的场景模板。")

    has_bound_subjects = any(entry.subjects for entry in scenes)
    if not has_bound_subjects and not subjects:
        raise TemplateRandomError("模板库里没有可用的主体模板。")

    return TemplateLibrary(layout_templates=layouts, scene_templates=scenes, subject_templates=subjects)


def pick_subject(scene_entry: SceneTemplateEntry, global_subjects: list[str]) -> str:
    pool = scene_entry.subjects or global_subjects
    if not pool:
        raise TemplateRandomError(f"场景模板未绑定可用主体：{scene_entry.content}")
    return random.choice(pool)


def pick_templates(library: TemplateLibrary, count: int, unique_scene: bool) -> list[SelectedTemplate]:
    if not unique_scene:
        return [
            SelectedTemplate(
                layout_template=random.choice(library.layout_templates),
                scene_template=scene_entry.content,
                subject=pick_subject(scene_entry, library.subject_templates),
            )
            for scene_entry in (random.choice(library.scene_templates) for _ in range(count))
        ]

    selected: list[SelectedTemplate] = []
    scene_sequence: list[SceneTemplateEntry] = []
    while len(scene_sequence) < count:
        round_scenes = list(library.scene_templates)
        random.shuffle(round_scenes)
        scene_sequence.extend(round_scenes)

    for scene_entry in scene_sequence[:count]:
        selected.append(
            SelectedTemplate(
                layout_template=random.choice(library.layout_templates),
                scene_template=scene_entry.content,
                subject=pick_subject(scene_entry, library.subject_templates),
            )
        )

    return selected


def format_color_option(color_option: tuple[str, str]) -> str:
    color_name, color_hex = color_option
    return f"{color_name} {color_hex}"


def build_color_replacements(count: int) -> list[str]:
    if count <= 0:
        return []

    color_pool = list(COLOR_OPTIONS)
    random.shuffle(color_pool)
    replacements: list[str] = []

    while len(replacements) < count:
        if not color_pool:
            color_pool = list(COLOR_OPTIONS)
            random.shuffle(color_pool)
        replacements.append(format_color_option(color_pool.pop()))

    return replacements


def has_color_placeholder_before(prompt: str, color_placeholder: str, subject_placeholder: str) -> bool:
    subject_color_pattern = re.compile(
        rf"{re.escape(color_placeholder)}(?=\s*{re.escape(subject_placeholder)})"
    )
    return bool(subject_color_pattern.search(prompt))


def replace_color_placeholders_before(
    prompt: str,
    color_placeholder: str,
    subject_placeholder: str,
    color_text: str | None = None,
) -> str:
    subject_color_pattern = re.compile(
        rf"{re.escape(color_placeholder)}(?=\s*{re.escape(subject_placeholder)})"
    )
    if not subject_color_pattern.search(prompt):
        return prompt

    if color_text is None:
        color_text = format_color_option(random.choice(COLOR_OPTIONS))

    return subject_color_pattern.sub(color_text, prompt)


def render_subject_text(subject_template: str, color_text: str | None = None, strip_embedded_color: bool = False) -> str:
    rendered = subject_template
    if COLOR_PLACEHOLDER not in rendered:
        return rendered.strip()

    if strip_embedded_color:
        return rendered.replace(COLOR_PLACEHOLDER, "", 1).strip()

    if color_text is None:
        color_text = format_color_option(random.choice(COLOR_OPTIONS))

    return rendered.replace(COLOR_PLACEHOLDER, color_text).strip()


def build_all_colors_text() -> str:
    return " / ".join(format_color_option(option) for option in COLOR_OPTIONS)


def build_all_subjects_text(global_subjects: list[str], fixed_color: str | None = None) -> str:
    color_values = build_color_replacements(len(global_subjects))
    rendered_subjects: list[str] = []

    for index, subject_template in enumerate(global_subjects):
        color_text = fixed_color if fixed_color is not None else color_values[index]
        rendered_subjects.append(render_subject_text(subject_template, color_text))

    return " / ".join(rendered_subjects)


def replace_color_placeholders(prompt: str) -> str:
    color_count = prompt.count(COLOR_PLACEHOLDER)
    if color_count == 0:
        return prompt

    for replacement in build_color_replacements(color_count):
        prompt = prompt.replace(COLOR_PLACEHOLDER, replacement, 1)

    return prompt


def replace_color_subject_pairs(
    prompt: str,
    subject_templates: list[str],
    color_placeholder: str,
    subject_placeholder: str,
    fixed_color: str | None = None,
    fixed_subject_template: str | None = None,
) -> str:
    pair_pattern = re.compile(
        rf"{re.escape(color_placeholder)}\s*{re.escape(subject_placeholder)}"
    )
    if not pair_pattern.search(prompt):
        return prompt

    color_pool = build_color_replacements(prompt.count(color_placeholder))
    subject_pool = list(subject_templates)
    random.shuffle(subject_pool)

    def replace_pair(_match: re.Match[str]) -> str:
        color_text = fixed_color if fixed_color is not None else color_pool.pop(0)
        if fixed_subject_template is not None:
            subject_template = fixed_subject_template
        elif subject_templates:
            nonlocal subject_pool
            if not subject_pool:
                subject_pool = list(subject_templates)
                random.shuffle(subject_pool)
            subject_template = subject_pool.pop()
        else:
            subject_template = ""
        subject_text = render_subject_text(subject_template, color_text, strip_embedded_color=True)
        return f"{color_text}{subject_text}"

    return pair_pattern.sub(replace_pair, prompt)


def render_prompt(template: SelectedTemplate, library: TemplateLibrary) -> str:
    prompt = template.layout_template
    subject_uses_prompt_color = has_color_placeholder_before(prompt, COLOR_PLACEHOLDER, SUBJECT_PLACEHOLDER)
    subject_color = None
    if COLOR_PLACEHOLDER in template.subject or subject_uses_prompt_color:
        subject_color = format_color_option(random.choice(COLOR_OPTIONS))
    subject_text = render_subject_text(
        template.subject,
        subject_color,
        strip_embedded_color=subject_uses_prompt_color,
    )

    fixed_color_required = FIXED_COLOR_PLACEHOLDER in prompt or FIXED_SUBJECT_PLACEHOLDER in prompt
    fixed_color = format_color_option(random.choice(COLOR_OPTIONS)) if fixed_color_required else None

    fixed_subject_template = random.choice(library.subject_templates) if library.subject_templates else ""
    fixed_subject_uses_prompt_color = has_color_placeholder_before(
        prompt,
        FIXED_COLOR_PLACEHOLDER,
        FIXED_SUBJECT_PLACEHOLDER,
    )
    fixed_subject_text = render_subject_text(
        fixed_subject_template,
        fixed_color,
        strip_embedded_color=fixed_subject_uses_prompt_color,
    )
    all_subjects_text = build_all_subjects_text(library.subject_templates)

    prompt = prompt.replace(SCENE_PLACEHOLDER, template.scene_template)
    prompt = replace_color_subject_pairs(prompt, library.subject_templates, COLOR_PLACEHOLDER, SUBJECT_PLACEHOLDER)
    prompt = replace_color_subject_pairs(
        prompt,
        library.subject_templates,
        FIXED_COLOR_PLACEHOLDER,
        FIXED_SUBJECT_PLACEHOLDER,
        fixed_color,
        fixed_subject_template,
    )
    prompt = replace_color_placeholders_before(prompt, COLOR_PLACEHOLDER, SUBJECT_PLACEHOLDER, subject_color)
    prompt = replace_color_placeholders_before(
        prompt,
        FIXED_COLOR_PLACEHOLDER,
        FIXED_SUBJECT_PLACEHOLDER,
        fixed_color,
    )
    prompt = prompt.replace(SUBJECT_PLACEHOLDER, subject_text)
    prompt = prompt.replace(FIXED_SUBJECT_PLACEHOLDER, fixed_subject_text)
    prompt = prompt.replace(ALL_SUBJECTS_PLACEHOLDER, all_subjects_text)
    if fixed_color is not None:
        prompt = prompt.replace(FIXED_COLOR_PLACEHOLDER, fixed_color)
    prompt = prompt.replace(ALL_COLORS_PLACEHOLDER, build_all_colors_text())
    prompt = replace_color_placeholders(prompt)
    return prompt.strip()


def build_unique_filename(batch_timestamp: str, index: int, total_count: int) -> str:
    if total_count <= 1:
        return f"{batch_timestamp}_photo"
    return f"{batch_timestamp}_{index:02d}_photo"


def decode_subprocess_output(data: bytes | None) -> str:
    if not data:
        return ""
    for encoding in ("utf-8", "gbk", "cp936"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def is_retryable_generation_error(error_text: str) -> bool:
    lowered = error_text.lower()
    retryable_markers = (
        "temporarily unavailable",
        "timeout",
        "timed out",
        "connection reset",
        "connection aborted",
        "recv failure",
        "curl: (56)",
        "502",
        "503",
        "504",
        "524",
    )
    return any(marker in lowered for marker in retryable_markers)


def run_image_generation(prompt: str, output_dir: str, filename: str, image2_script: str | None) -> str:
    script_path = image2_script or str((Path(__file__).resolve().parents[1] / "image2-generate" / "scripts" / "generate_image.py").resolve())
    command = [
        sys.executable,
        script_path,
        "--prompt",
        prompt,
        "--filename",
        filename,
        "--output-dir",
        output_dir,
    ]

    last_error = ""
    for attempt in range(1, 4):
        completed = subprocess.run(command, capture_output=True, text=False, check=False)
        stdout_text = decode_subprocess_output(completed.stdout)
        stderr_text = decode_subprocess_output(completed.stderr)

        if completed.returncode == 0:
            lines = [line.strip() for line in stdout_text.splitlines() if line.strip()]
            if not lines:
                raise TemplateRandomError("生图脚本没有返回图片路径。")
            return lines[-1]

        last_error = (stderr_text or stdout_text).strip()
        if attempt < 3 and is_retryable_generation_error(last_error):
            time.sleep(10 * attempt)
            continue
        break

    raise TemplateRandomError(f"生图失败：{last_error}")


def generate_one(index: int, prompt: str, output_dir: str, batch_timestamp: str, total_count: int, image2_script: str | None) -> dict:
    filename = build_unique_filename(batch_timestamp, index, total_count)
    image_path = run_image_generation(prompt, output_dir, filename, image2_script)
    return {
        "index": index,
        "prompt": prompt,
        "filename": filename,
        "image_path": image_path,
    }


def generate_images(prompts: list[str], output_dir: str, concurrency: int, image2_script: str | None) -> list[dict]:
    batch_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    total_count = len(prompts)
    results: dict[int, dict] = {}
    errors: list[str] = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=concurrency) as executor:
        future_map = {
            executor.submit(generate_one, index, prompt, output_dir, batch_timestamp, total_count, image2_script): index
            for index, prompt in enumerate(prompts, start=1)
        }
        for future in concurrent.futures.as_completed(future_map):
            index = future_map[future]
            try:
                results[index] = future.result()
            except Exception as exc:
                errors.append(f"#{index}: {exc}")

    if errors:
        raise TemplateRandomError("\n".join(errors))

    return [results[index] for index in range(1, total_count + 1)]


def main() -> int:
    args = parse_args()

    if args.seed is not None:
        random.seed(args.seed)

    count = max(1, args.count)
    concurrency = max(1, min(args.concurrency, count))
    output_dir = str(Path(args.output_dir))

    try:
        library = load_template_library(Path(args.template_path))
        selected_templates = pick_templates(library, count, args.unique_scene)
        prompts = [render_prompt(item, library) for item in selected_templates]

        if args.prompts_only:
            print(
                json.dumps(
                    {
                        "mode": "prompts_only",
                        "outputDirectory": output_dir,
                        "prompts": prompts,
                    },
                    ensure_ascii=False,
                )
            )
            return 0

        results = generate_images(prompts, output_dir, concurrency, args.image2_script)
        print(
            json.dumps(
                {
                    "mode": "generated",
                    "outputDirectory": output_dir,
                    "prompts": prompts,
                    "results": results,
                },
                ensure_ascii=False,
            )
        )
        return 0
    except TemplateRandomError as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
