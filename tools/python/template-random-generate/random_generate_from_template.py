#!/usr/bin/env python3
"""Template-based random prompt generation and optional image generation."""

from __future__ import annotations

import argparse
import concurrent.futures
import json
import random
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


DEFAULT_TEMPLATE_PATH = Path(r"D:\temu_auto\temp\文生图模板库_Codex.xlsx")
DEFAULT_OUTPUT_DIR = Path(r"D:\temu_auto\review")
SHEET_NAME = "文生图模板库"
LAYOUT_COLUMN = "布局模板"
SCENE_COLUMN = "场景模板"
SUBJECT_COLUMN = "主体"

COLOR_NAMES = (
    "黑色",
    "米白色",
    "深棕色",
    "浅灰色",
    "酒红色",
    "宝蓝色",
)


class TemplateRandomError(Exception):
    """Raised when template generation fails with a user-facing message."""


@dataclass
class TemplateRow:
    scene_options: list[str]
    subject_options: list[str]


@dataclass
class SelectedTemplate:
    layout_template: str
    scene_template: str
    subject: str


@dataclass
class TemplateLibrary:
    layout_templates: list[str]
    rows: list[TemplateRow]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate prompts from the local Excel template library.")
    parser.add_argument("--template-path", default=str(DEFAULT_TEMPLATE_PATH))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--count", type=int, default=5)
    parser.add_argument("--concurrency", type=int, default=2)
    parser.add_argument("--image2-script")
    parser.add_argument("--seed", type=int)
    parser.add_argument("--unique-scene", action="store_true")
    parser.add_argument("--prompts-only", action="store_true")
    return parser.parse_args()


def split_variants(value: str | None) -> list[str]:
    if value is None:
        return []

    text = str(value).replace("\r\n", "\n").replace("\r", "\n").replace("，", "/")
    parts = re.split(r"[\/\n]+", text)
    return [part.strip(" -\t") for part in parts if part and part.strip(" -\t")]


def load_template_library(template_path: Path) -> TemplateLibrary:
    if not template_path.exists():
        raise TemplateRandomError(f"模板库不存在：{template_path}")

    workbook = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise TemplateRandomError(f"模板库中缺少工作表：{SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise TemplateRandomError("模板库没有可用数据。")

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    try:
        layout_index = header.index(LAYOUT_COLUMN)
        scene_index = header.index(SCENE_COLUMN)
        subject_index = header.index(SUBJECT_COLUMN)
    except ValueError as exc:
        raise TemplateRandomError("模板库表头缺少“布局模板 / 场景模板 / 主体”列。") from exc

    layout_templates: list[str] = []
    template_rows: list[TemplateRow] = []

    for raw_row in rows[1:]:
        if not raw_row:
            continue

        layout = raw_row[layout_index] if len(raw_row) > layout_index else None
        scene = raw_row[scene_index] if len(raw_row) > scene_index else None
        subject = raw_row[subject_index] if len(raw_row) > subject_index else None

        if layout and str(layout).strip():
            layout_templates.append(str(layout).strip())

        scenes = split_variants(scene)
        subjects = split_variants(subject)
        if scenes and subjects:
            template_rows.append(TemplateRow(scene_options=scenes, subject_options=subjects))

    if not layout_templates:
        raise TemplateRandomError("模板库里没有可用的布局模板。")
    if not template_rows:
        raise TemplateRandomError("模板库里没有可用的场景模板和主体。")

    return TemplateLibrary(layout_templates=layout_templates, rows=template_rows)


def choose_selected_template(layout_template: str, row: TemplateRow, scene_override: str | None = None) -> SelectedTemplate:
    scene_template = scene_override if scene_override is not None else random.choice(row.scene_options)
    subject = random.choice(row.subject_options)
    return SelectedTemplate(
        layout_template=layout_template,
        scene_template=scene_template,
        subject=subject,
    )


def pick_templates(library: TemplateLibrary, count: int, unique_scene: bool) -> list[SelectedTemplate]:
    if not unique_scene:
        return [
            choose_selected_template(
                random.choice(library.layout_templates),
                random.choice(library.rows),
            )
            for _ in range(count)
        ]

    scene_map: dict[str, list[TemplateRow]] = {}
    for row in library.rows:
        for scene in row.scene_options:
            scene_map.setdefault(scene, []).append(row)

    unique_scenes = list(scene_map.keys())
    if not unique_scenes:
        raise TemplateRandomError("模板库里没有可用的场景模板。")

    selected: list[SelectedTemplate] = []
    scene_sequence: list[str] = []
    while len(scene_sequence) < count:
        round_scenes = unique_scenes[:]
        random.shuffle(round_scenes)
        scene_sequence.extend(round_scenes)

    for scene in scene_sequence[:count]:
        selected.append(
            choose_selected_template(
                random.choice(library.layout_templates),
                random.choice(scene_map[scene]),
                scene_override=scene,
            )
        )
    return selected


def render_prompt(template: SelectedTemplate) -> str:
    prompt = template.layout_template
    prompt = prompt.replace("{场景模板}", template.scene_template)
    prompt = prompt.replace("{主体}", template.subject)
    prompt = prompt.replace("{颜色}", random.choice(COLOR_NAMES))
    return prompt.strip()


def build_unique_filename(batch_timestamp: str, index: int, total_count: int) -> str:
    if total_count <= 1:
        return f"{batch_timestamp}_photo"
    return f"{batch_timestamp}_{index:02d}_photo"


def decode_subprocess_output(data: bytes | None) -> str:
    if not data:
        return ""
    for encoding in ("utf-8", "gbk", "cp936"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def is_retryable_generation_error(error_text: str) -> bool:
    lowered = error_text.lower()
    retryable_markers = (
        "temporarily unavailable",
        "timeout",
        "timed out",
        "502",
        "503",
        "504",
        "524",
    )
    return any(marker in lowered for marker in retryable_markers)


def run_image_generation(prompt: str, output_dir: str, filename: str, image2_script: str | None) -> str:
    script_path = image2_script or str((Path(__file__).resolve().parents[1] / "image2-generate" / "scripts" / "generate_image.py").resolve())
    command = [
        sys.executable,
        script_path,
        "--prompt",
        prompt,
        "--filename",
        filename,
        "--output-dir",
        output_dir,
    ]

    last_error = ""
    for attempt in range(1, 4):
        completed = subprocess.run(command, capture_output=True, text=False, check=False)
        stdout_text = decode_subprocess_output(completed.stdout)
        stderr_text = decode_subprocess_output(completed.stderr)

        if completed.returncode == 0:
            lines = [line.strip() for line in stdout_text.splitlines() if line.strip()]
            if not lines:
                raise TemplateRandomError("生图脚本没有返回图片路径。")
            return lines[-1]

        last_error = (stderr_text or stdout_text).strip()
        if attempt < 3 and is_retryable_generation_error(last_error):
            time.sleep(10 * attempt)
            continue
        break

    raise TemplateRandomError(f"生图失败：{last_error}")


def generate_one(index: int, prompt: str, output_dir: str, batch_timestamp: str, total_count: int, image2_script: str | None) -> dict:
    filename = build_unique_filename(batch_timestamp, index, total_count)
    image_path = run_image_generation(prompt, output_dir, filename, image2_script)
    return {
        "index": index,
        "prompt": prompt,
        "filename": filename,
        "image_path": image_path,
    }


def generate_images(prompts: list[str], output_dir: str, concurrency: int, image2_script: str | None) -> list[dict]:
    batch_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    total_count = len(prompts)
    results: dict[int, dict] = {}
    errors: list[str] = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=concurrency) as executor:
        future_map = {
            executor.submit(generate_one, index, prompt, output_dir, batch_timestamp, total_count, image2_script): index
            for index, prompt in enumerate(prompts, start=1)
        }
        for future in concurrent.futures.as_completed(future_map):
            index = future_map[future]
            try:
                results[index] = future.result()
            except Exception as exc:
                errors.append(f"#{index}: {exc}")

    if errors:
        raise TemplateRandomError("；".join(errors))

    return [results[index] for index in range(1, total_count + 1)]


def main() -> int:
    args = parse_args()

    if args.seed is not None:
        random.seed(args.seed)

    count = max(1, args.count)
    concurrency = max(1, min(args.concurrency, count))
    output_dir = str(Path(args.output_dir))

    try:
        library = load_template_library(Path(args.template_path))
        selected_templates = pick_templates(library, count, args.unique_scene)
        prompts = [render_prompt(item) for item in selected_templates]

        if args.prompts_only:
            print(
                json.dumps(
                    {
                        "mode": "prompts_only",
                        "outputDirectory": output_dir,
                        "prompts": prompts,
                    },
                    ensure_ascii=False,
                )
            )
            return 0

        results = generate_images(prompts, output_dir, concurrency, args.image2_script)
        print(
            json.dumps(
                {
                    "mode": "generated",
                    "outputDirectory": output_dir,
                    "prompts": prompts,
                    "results": results,
                },
                ensure_ascii=False,
            )
        )
        return 0
    except TemplateRandomError as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
