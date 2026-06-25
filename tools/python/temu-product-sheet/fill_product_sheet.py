import argparse
from copy import copy
import json
import os
import random
import re
import subprocess
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INDEX = Path("D:/temu_auto/json/\u5c3a\u5bf8\u89c4\u683c\u7d22\u5f15.json")
DEFAULT_SOURCE = Path("D:/temu_auto/\u5c3a\u5bf8\u89c4\u683c.xlsx")
DEFAULT_TEMPLATE = Path("D:/temu_auto/\u5546\u54c1\u4fe1\u606f\u8868.xlsx")
DEFAULT_OUTPUT_DIR = Path("D:/temu_auto/excel")
DEFAULT_ASSERT_DIR = Path("D:/temu_auto/assert")
DEFAULT_TITLE_JSON = Path("D:/temu_auto/json/title.json")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate a dated Temu product sheet workbook from SKU sizes."
    )
    parser.add_argument(
        "--sizes",
        nargs="+",
        help='One or more SKU sizes such as "60*168cm".',
    )
    parser.add_argument(
        "--sp-dir",
        default=None,
        help="SPxx folder path. When provided, auto-extract sizes from main\\2-尺寸.png.",
    )
    parser.add_argument(
        "--assert-dir",
        default=str(DEFAULT_ASSERT_DIR),
        help="Root folder containing SPxx directories. Used by default when --sizes and --sp-dir are not provided.",
    )
    parser.add_argument(
        "--product-id",
        default="SP1",
        help="Value to write into 商品编号 for matched rows.",
    )
    parser.add_argument(
        "--template",
        default=str(DEFAULT_TEMPLATE),
        help="Path to the product sheet template workbook.",
    )
    parser.add_argument(
        "--index",
        default=str(DEFAULT_INDEX),
        help="Path to the JSON size index.",
    )
    parser.add_argument(
        "--source",
        default=str(DEFAULT_SOURCE),
        help="Path to the source spec workbook used when rebuilding the index.",
    )
    parser.add_argument(
        "--title-json",
        default=str(DEFAULT_TITLE_JSON),
        help="Path to the title JSON used to randomize 产品标题 for each SPxx row.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Directory for the dated output workbook.",
    )
    parser.add_argument(
        "--date",
        default=date.today().isoformat(),
        help="Date string for the output filename, format YYYY-MM-DD.",
    )
    parser.add_argument(
        "--output-name",
        default=None,
        help="Optional explicit output filename.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=None,
        help="Optional random seed for repeatable price generation.",
    )
    return parser.parse_args()


def ensure_index(index_path, source_path):
    if index_path.exists() and source_path.exists():
        if index_path.stat().st_mtime >= source_path.stat().st_mtime:
            return
    elif index_path.exists():
        return

    build_script = Path(__file__).with_name("build_size_index.py")
    cmd = [
        sys.executable,
        str(build_script),
        "--source",
        str(source_path),
        "--output",
        str(index_path),
    ]
    subprocess.run(cmd, check=True)


def extract_sizes_from_sp_dir(sp_dir):
    main_dir = Path(sp_dir) / "main"
    candidates = sorted(main_dir.glob("2-*.png"))
    if not candidates:
        raise FileNotFoundError(f"Automatic size image not found under: {main_dir}")

    ps_script = f"""
$ErrorActionPreference = 'Stop'
Add-Type -AssemblyName System.Runtime.WindowsRuntime
$null = [Windows.Storage.StorageFile, Windows.Storage, ContentType = WindowsRuntime]
$null = [Windows.Media.Ocr.OcrEngine, Windows.Foundation, ContentType = WindowsRuntime]
$null = [Windows.Graphics.Imaging.BitmapDecoder, Windows.Foundation, ContentType = WindowsRuntime]
$null = [Windows.Storage.Streams.IRandomAccessStream, Windows.Foundation, ContentType = WindowsRuntime]
$null = [Windows.Graphics.Imaging.SoftwareBitmap, Windows.Foundation, ContentType = WindowsRuntime]
$null = [Windows.Media.Ocr.OcrResult, Windows.Foundation, ContentType = WindowsRuntime]

function AwaitResult($op, [Type]$resultType) {{
  $method = [System.WindowsRuntimeSystemExtensions].GetMethods() | Where-Object {{
    $_.Name -eq 'AsTask' -and $_.IsGenericMethod -and $_.GetParameters().Count -eq 1
  }} | Select-Object -First 1
  $generic = $method.MakeGenericMethod($resultType)
  $task = $generic.Invoke($null, @($op))
  $task.Wait()
  $task.Result
}}

$mainDir = '{str(main_dir).replace("'", "''")}'
$image = Get-ChildItem -LiteralPath $mainDir -Filter '2-*.png' | Select-Object -First 1
if (-not $image) {{
  throw "No 2-*.png file found under $mainDir"
}}
$file = AwaitResult ([Windows.Storage.StorageFile]::GetFileFromPathAsync($image.FullName)) ([Windows.Storage.StorageFile])
$stream = AwaitResult ($file.OpenAsync([Windows.Storage.FileAccessMode]::Read)) ([Windows.Storage.Streams.IRandomAccessStream])
$decoder = AwaitResult ([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($stream)) ([Windows.Graphics.Imaging.BitmapDecoder])
$bitmap = AwaitResult ($decoder.GetSoftwareBitmapAsync()) ([Windows.Graphics.Imaging.SoftwareBitmap])
$engine = [Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages()
$result = AwaitResult ($engine.RecognizeAsync($bitmap)) ([Windows.Media.Ocr.OcrResult])
$result.Text
"""
    completed = subprocess.run(
        ["powershell", "-NoProfile", "-Command", "[Console]::OutputEncoding = [System.Text.Encoding]::UTF8; " + ps_script],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    return parse_sizes_from_ocr_text(completed.stdout)


def parse_sizes_from_ocr_text(text):
    normalized = text.replace("\r", "\n")
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*[*xX]\s*(\d+(?:\.\d+)?)\s*cm", normalized, re.I)
    unique = []
    seen = set()
    for width_text, length_text in matches:
        size_text = f"{int(float(width_text))}*{int(float(length_text))}cm"
        if size_text in seen:
            continue
        seen.add(size_text)
        unique.append(size_text)
    return unique


def parse_size(size_text):
    match = re.fullmatch(r"\s*(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)\s*cm\s*", size_text, re.I)
    if not match:
        raise ValueError(f"Unsupported size format: {size_text}")
    width = int(float(match.group(1)))
    length = int(float(match.group(2)))
    return width, length


def load_records(index_path):
    with index_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    return payload["records"]


def load_titles(title_json_path):
    with Path(title_json_path).open("r", encoding="utf-8") as handle:
        payload = json.load(handle)

    titles = payload.get("titles")
    if not isinstance(titles, list):
        raise ValueError(f"Invalid title JSON format: {title_json_path}")

    cleaned = []
    for item in titles:
        if isinstance(item, str):
            text = item.strip()
            if text:
                cleaned.append(text)

    if not cleaned:
        raise ValueError(f"No usable titles found in: {title_json_path}")
    return cleaned


def match_record(records, width, length):
    for record in records:
        if record["width_cm"] != width:
            continue
        if record["length_min_cm"] is None or record["length_max_cm"] is None:
            continue
        if record["length_min_cm"] <= length <= record["length_max_cm"]:
            return record
    return None


def random_price(record):
    price_min = record["declared_price_min"]
    price_max = record["declared_price_max"]
    if price_min is None or price_max is None:
        return None
    min_cents = int(round(price_min * 100))
    max_cents = int(round(price_max * 100))
    if max_cents < min_cents:
        raise ValueError(f"Invalid price range: {record['declared_price_range_text']}")
    return random.randint(min_cents, max_cents) / 100.0


def clear_sheet_rows(sheet, start_row, end_col):
    max_row = max(sheet.max_row, start_row)
    for row in range(start_row, max_row + 1):
        for col in range(1, end_col + 1):
            sheet.cell(row, col).value = None


def clear_rows(sheet, start_row=2, end_row=500, end_col=7):
    for row in range(start_row, end_row + 1):
        for col in range(1, end_col + 1):
            sheet.cell(row, col).value = None


def apply_template_row_format(sheet, target_row, template_row, start_col, end_col):
    source_height = sheet.row_dimensions[template_row].height
    if source_height is not None:
        sheet.row_dimensions[target_row].height = source_height

    for col in range(start_col, end_col + 1):
        source = sheet.cell(template_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def write_main_rows(sheet, main_rows):
    for row_index, item in enumerate(main_rows, start=2):
        apply_template_row_format(sheet, row_index, 2, 1, 5)
        sheet.cell(row_index, 1).value = item["product_id"]
        sheet.cell(row_index, 2).value = item["title"]
        sheet.cell(row_index, 3).value = item["main_path"]
        sheet.cell(row_index, 4).value = item["detail_path"]
        sheet.cell(row_index, 5).value = item["sku_path"]


def write_rows(sheet, matched_rows):
    for row_index, item in enumerate(matched_rows, start=2):
        apply_template_row_format(sheet, row_index, 2, 1, 7)
        record = item["record"]
        price = item["price"]
        sheet.cell(row_index, 1).value = item["product_id"]
        sheet.cell(row_index, 2).value = item["size_text"]
        sheet.cell(row_index, 3).value = record["longest_edge_cm"]
        sheet.cell(row_index, 4).value = record["second_longest_edge_cm"]
        sheet.cell(row_index, 5).value = record["shortest_edge_cm"]
        sheet.cell(row_index, 6).value = record["weight_g"]
        sheet.cell(row_index, 7).value = price
        sheet.cell(row_index, 7).number_format = "0.00"


def collect_sp_directories(assert_dir):
    assert_root = Path(assert_dir)
    if not assert_root.exists():
        raise FileNotFoundError(f"Assert root not found: {assert_root}")
    if not assert_root.is_dir():
        raise NotADirectoryError(f"Assert root is not a directory: {assert_root}")

    sp_dirs = [
        path for path in sorted(assert_root.iterdir())
        if path.is_dir() and re.fullmatch(r"SP\d+", path.name, re.I)
    ]
    if not sp_dirs:
        raise FileNotFoundError(f"No SPxx folders found under: {assert_root}")
    return sp_dirs


def build_main_row(sp_dir, workbook, titles):
    title = random.choice(titles)
    return {
        "product_id": sp_dir.name,
        "title": title,
        "main_path": str((sp_dir / "main").resolve()),
        "detail_path": str((sp_dir / "detail").resolve()),
        "sku_path": str((sp_dir / "sku").resolve()),
    }


def process_sp_dir(sp_dir, records):
    size_texts = extract_sizes_from_sp_dir(sp_dir)
    if not size_texts:
        raise ValueError(f"No sizes were extracted from: {sp_dir}")

    matched_rows = []
    skipped_sizes = []
    for size_text in size_texts:
        width, length = parse_size(size_text)
        record = match_record(records, width, length)
        if record is None:
            skipped_sizes.append(size_text)
            continue
        matched_rows.append(
            {
                "product_id": sp_dir.name,
                "size_text": size_text,
                "record": record,
                "price": random_price(record),
            }
        )

    return {
        "product_id": sp_dir.name,
        "size_texts": size_texts,
        "matched_rows": matched_rows,
        "skipped_sizes": skipped_sizes,
    }


def output_path_for(args):
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if args.output_name:
        return output_dir / args.output_name
    return output_dir / "\u5546\u54c1\u4fe1\u606f\u8868.xlsx"


def main():
    args = parse_args()
    if args.seed is not None:
        random.seed(args.seed)

    template_path = Path(args.template)
    index_path = Path(args.index)
    source_path = Path(args.source)
    title_json_path = Path(args.title_json)
    ensure_index(index_path, source_path)
    records = load_records(index_path)
    titles = load_titles(title_json_path)

    workbook = load_workbook(template_path)
    main_sheet = workbook.worksheets[0]
    sku_sheet = workbook.worksheets[1]

    main_rows = []
    matched_rows = []
    summaries = []

    if args.sizes:
        size_texts = list(args.sizes)
        current_matched_rows = []
        skipped_sizes = []
        for size_text in size_texts:
            width, length = parse_size(size_text)
            record = match_record(records, width, length)
            if record is None:
                skipped_sizes.append(size_text)
                continue
            current_matched_rows.append(
                {
                    "product_id": args.product_id,
                    "size_text": size_text,
                    "record": record,
                    "price": random_price(record),
                }
            )

        matched_rows.extend(current_matched_rows)
        summaries.append(
            {
                "product_id": args.product_id,
                "size_texts": size_texts,
                "matched_rows": current_matched_rows,
                "skipped_sizes": skipped_sizes,
            }
        )
    elif args.sp_dir:
        sp_path = Path(args.sp_dir)
        main_rows.append(build_main_row(sp_path, workbook, titles))
        summary = process_sp_dir(sp_path, records)
        matched_rows.extend(summary["matched_rows"])
        summaries.append(summary)
    else:
        for sp_path in collect_sp_directories(args.assert_dir):
            main_rows.append(build_main_row(sp_path, workbook, titles))
            summaries.append(process_sp_dir(sp_path, records))
        for summary in summaries:
            matched_rows.extend(summary["matched_rows"])

    clear_sheet_rows(main_sheet, start_row=2, end_col=6)
    clear_sheet_rows(sku_sheet, start_row=2, end_col=7)
    if main_rows:
        write_main_rows(main_sheet, main_rows)
    write_rows(sku_sheet, matched_rows)

    output_path = output_path_for(args)
    workbook.save(output_path)

    print(output_path)
    print(f"products={len(main_rows) if main_rows else (1 if summaries else 0)}")
    print(f"matched={len(matched_rows)}")
    skipped_total = 0
    for summary in summaries:
        skipped_total += len(summary["skipped_sizes"])
        print("product={0};sizes={1};matched={2};skipped={3}".format(
            summary["product_id"],
            ",".join(summary["size_texts"]),
            len(summary["matched_rows"]),
            len(summary["skipped_sizes"]),
        ))
        if summary["skipped_sizes"]:
            print("skipped_sizes[{0}]={1}".format(summary["product_id"], ",".join(summary["skipped_sizes"])))
    print(f"skipped={skipped_total}")


if __name__ == "__main__":
    main()
